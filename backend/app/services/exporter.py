"""Excel export service for schedule data."""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models.schemas import (
    Employee,
    ShiftType,
    DailySchedule,
)

# 模板样式（按用户给的 Excel 解析）
# 白班：黄；大夜：蓝；小夜：白底；睡觉：白底；其他：绿
SHIFT_FILLS = {
    "DAY": PatternFill(start_color="FFFED961", end_color="FFFED961", fill_type="solid"),
    "LATE_NIGHT": PatternFill(start_color="FF5B9BD5", end_color="FF5B9BD5", fill_type="solid"),
    "VACATION": PatternFill(start_color="FF75BD42", end_color="FF75BD42", fill_type="solid"),
    "OTHER": PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid"),
    "NONE": PatternFill(fill_type=None),
}

CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _normalize_shift_type(value: str | ShiftType | None) -> str:
    if value is None:
        return "NONE"
    if isinstance(value, ShiftType):
        return value.value
    return str(value)


def _get_display_text(shift_type: str, label: str | None) -> str:
    if label:
        return label

    mapping = {
        "DAY": "白班",
        "SLEEP": "睡觉",
        "MINI_NIGHT": "小夜",
        "LATE_NIGHT": "大夜",
        "VACATION": "休假",
        "CUSTOM": "其他",
        "NONE": "",
    }
    return mapping.get(shift_type, "其他")


def _get_fill_by_value(shift_type: str, text: str) -> PatternFill:
    if shift_type == "DAY":
        return SHIFT_FILLS["DAY"]
    if shift_type == "LATE_NIGHT":
        return SHIFT_FILLS["LATE_NIGHT"]
    if shift_type in ("VACATION",):
        return SHIFT_FILLS["VACATION"]
    if shift_type in ("SLEEP", "MINI_NIGHT", "NONE"):
        return SHIFT_FILLS["NONE"]

    if "白班" in text:
        return SHIFT_FILLS["DAY"]
    if "大夜" in text:
        return SHIFT_FILLS["LATE_NIGHT"]
    if "休" in text or "假" in text or "培训" in text or "出差" in text:
        return SHIFT_FILLS["VACATION"]
    if "小夜" in text or "睡" in text:
        return SHIFT_FILLS["NONE"]
    return SHIFT_FILLS["OTHER"]


def _apply_row_base_style(ws, row_idx: int, max_col: int = 19):
    for col in range(1, max_col + 1):
        c = ws.cell(row=row_idx, column=col)
        c.alignment = CENTER_ALIGN
        c.border = THIN_BORDER
        c.font = Font(name="宋体", size=11, color="FF000000")


def _write_month_block(
    ws,
    start_row: int,
    month: str,
    group_id: str,
    schedules: list[DailySchedule],
    employees: list[Employee],
) -> int:
    """在单个工作表中写入一个月份块，返回下一个可写入行号。"""

    # 月份标题行
    title_cell = ws.cell(row=start_row, column=1, value=f"{month} {group_id}组")
    title_cell.font = Font(name="宋体", size=12, bold=True)
    title_cell.alignment = CENTER_ALIGN

    header_row = start_row + 1

    # 表头布局：A=日期，B~G + I~S 员工列，H 为空白分隔列
    ws.cell(row=header_row, column=1, value="日期")

    left_count = min(6, len(employees))
    employee_columns: dict[str, int] = {}
    for idx, emp in enumerate(employees):
        if idx < left_count:
            col = 2 + idx  # B..G
        else:
            col = 9 + (idx - left_count)  # I..S
        employee_columns[emp.id] = col
        ws.cell(row=header_row, column=col, value=emp.name)

    _apply_row_base_style(ws, header_row)

    # 数据行
    row_idx = header_row + 1
    for schedule in schedules:
        try:
            mm = int(schedule.date[5:7])
            dd = int(schedule.date[8:10])
            date_text = f"{mm}.{dd}"
        except Exception:
            date_text = schedule.date

        ws.cell(row=row_idx, column=1, value=date_text)
        _apply_row_base_style(ws, row_idx)

        record_map = {r.employee_id: r for r in schedule.records}
        for emp in employees:
            col = employee_columns.get(emp.id)
            if col is None:
                continue

            record = record_map.get(emp.id)
            shift_type = _normalize_shift_type(record.shift_type if record else None)
            text = _get_display_text(shift_type, record.label if record else None)

            cell = ws.cell(row=row_idx, column=col, value=text)
            cell.fill = _get_fill_by_value(shift_type, text)

        row_idx += 1

    # 块后留一个空行，便于区分月份
    return row_idx + 1


def _prepare_sheet(ws):
    # 列宽：模板中主表每列约 13
    for col in range(1, 20):
        ws.column_dimensions[get_column_letter(col)].width = 13


def export_schedule_to_excel(
    month: str,
    group_id: str,
    schedules: list[DailySchedule],
    employees: list[Employee],
) -> io.BytesIO:
    """导出单月 Excel（模板样式）"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month}-{group_id}"[:31]
    _prepare_sheet(ws)

    _write_month_block(
        ws=ws,
        start_row=1,
        month=month,
        group_id=group_id,
        schedules=schedules,
        employees=employees,
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_multi_month_schedule_to_excel(
    group_id: str,
    month_schedules: dict[str, list[DailySchedule]],
    month_employees: dict[str, list[Employee]],
) -> io.BytesIO:
    """导出多月 Excel：单个 Sheet 纵向拼接（不新建多 Sheet）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "排班汇总"
    _prepare_sheet(ws)

    current_row = 1
    for month in sorted(month_schedules.keys()):
        current_row = _write_month_block(
            ws=ws,
            start_row=current_row,
            month=month,
            group_id=group_id,
            schedules=month_schedules.get(month, []),
            employees=month_employees.get(month, []),
        )

    if not month_schedules:
        ws.cell(row=1, column=1, value="无可导出数据")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
